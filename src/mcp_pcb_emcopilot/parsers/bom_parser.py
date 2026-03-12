"""BOM (Bill of Materials) parser for CSV and Excel formats.

Supports:
- CSV files with flexible column detection
- Excel files (.xlsx, .xls) via openpyxl
- Automatic component type inference from reference designators
"""

import csv
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional, TextIO

logger = logging.getLogger(__name__)


@dataclass
class ParsedBOMItem:
    """Single BOM line item."""
    line_number: int
    references: str  # Comma-separated: "R1,R2,R3"
    quantity: int
    value: Optional[str] = None
    part_number: Optional[str] = None
    manufacturer: Optional[str] = None
    description: Optional[str] = None
    footprint: Optional[str] = None
    supplier: Optional[str] = None
    supplier_part_number: Optional[str] = None
    unit_price: Optional[float] = None
    component_type: Optional[str] = None
    properties: dict = field(default_factory=dict)


@dataclass
class ParsedBOMData:
    """Complete parsed BOM data."""
    items: list[ParsedBOMItem] = field(default_factory=list)
    total_items: int = 0
    title: Optional[str] = None
    revision: Optional[str] = None
    warnings: list[str] = field(default_factory=list)


class BOMParser:
    """Parser for BOM files in CSV and Excel formats."""

    # Common column name variations
    COLUMN_MAPPINGS = {
        'reference': ['ref', 'reference', 'designator', 'references', 'ref des', 'part', 'component'],
        'quantity': ['qty', 'quantity', 'qnty', 'count', 'amount'],
        'value': ['value', 'val', 'component value'],
        'part_number': ['mpn', 'part number', 'part_number', 'mfg part', 'manufacturer part number', 'part #', 'p/n'],
        'manufacturer': ['mfr', 'manufacturer', 'mfg', 'vendor'],
        'description': ['description', 'desc', 'notes', 'comment'],
        'footprint': ['footprint', 'package', 'pkg'],
        'supplier': ['supplier', 'distributor', 'source'],
        'supplier_part_number': ['spn', 'supplier part', 'distributor part', 'supplier part number', 'digi-key part'],
        'unit_price': ['price', 'unit price', 'cost', 'unit_price', 'unit cost'],
    }

    def __init__(self):
        self.items: list[ParsedBOMItem] = []
        self.warnings: list[str] = []

    def parse(self, file_path: str) -> ParsedBOMData:
        """Parse BOM file with automatic format detection.

        Args:
            file_path: Path to BOM file (CSV or Excel)

        Returns:
            ParsedBOMData with line items

        Raises:
            ValueError: If file format is invalid or unsupported
        """
        logger.info(f"Parsing BOM file: {file_path}")

        path = Path(file_path)
        extension = path.suffix.lower()

        try:
            if extension == '.csv':
                return self._parse_csv(file_path)
            elif extension in ['.xlsx', '.xls']:
                return self._parse_excel(file_path)
            else:
                raise ValueError(f"Unsupported BOM format: {extension}")

        except Exception as e:
            logger.error(f"Failed to parse BOM: {e}")
            raise ValueError(f"BOM parse error: {str(e)}")

    def _parse_csv(self, file_path: str) -> ParsedBOMData:
        """Parse CSV BOM file with automatic delimiter detection."""
        # Detect delimiter
        with open(file_path, 'r', encoding='utf-8') as f:
            sample = f.read(4096)
            sniffer = csv.Sniffer()
            try:
                dialect = sniffer.sniff(sample)
                delimiter = dialect.delimiter
            except:
                delimiter = ','  # Default to comma

        # Parse CSV
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f, delimiter=delimiter)

            # Map column names to standardized keys
            column_map = self._map_columns(reader.fieldnames or [])

            for line_num, row in enumerate(reader, start=2):  # Start at 2 (after header)
                try:
                    item = self._parse_row(row, column_map, line_num)
                    if item:
                        self.items.append(item)
                except Exception as e:
                    self.warnings.append(f"Line {line_num}: {str(e)}")

        result = ParsedBOMData(
            items=self.items,
            total_items=len(self.items),
            warnings=self.warnings,
        )

        logger.info(f"Parsed {len(result.items)} BOM items from CSV")
        return result

    def _parse_excel(self, file_path: str) -> ParsedBOMData:
        """Parse Excel BOM file using openpyxl."""
        try:
            import openpyxl
        except ImportError:
            raise ValueError("openpyxl not installed. Install with: pip install openpyxl")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).lower().strip() if cell.value else '')

        # Map column names
        column_map = self._map_columns(headers)

        # Parse rows
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Convert row to dict
                row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                item = self._parse_row(row_dict, column_map, row_num)
                if item:
                    self.items.append(item)
            except Exception as e:
                self.warnings.append(f"Row {row_num}: {str(e)}")

        wb.close()

        result = ParsedBOMData(
            items=self.items,
            total_items=len(self.items),
            warnings=self.warnings,
        )

        logger.info(f"Parsed {len(result.items)} BOM items from Excel")
        return result

    def _map_columns(self, headers: list[str]) -> dict[str, str]:
        """Map file column names to standardized field names.

        Args:
            headers: List of column names from file

        Returns:
            Dictionary mapping file column name to standard field name
        """
        column_map = {}

        for header in headers:
            header_lower = header.lower().strip()

            # Try to match each standard field
            for field_name, variations in self.COLUMN_MAPPINGS.items():
                for variation in variations:
                    if variation in header_lower:
                        column_map[header] = field_name
                        break
                if header in column_map:
                    break

        return column_map

    def _parse_row(self, row: dict, column_map: dict, line_number: int) -> Optional[ParsedBOMItem]:
        """Parse a single BOM row.

        Args:
            row: Dictionary of column values
            column_map: Mapping of file columns to standard fields
            line_number: Line number for error reporting

        Returns:
            ParsedBOMItem or None if row should be skipped
        """
        # Extract mapped fields
        def get_field(field_name: str) -> Optional[str]:
            for col, mapped_field in column_map.items():
                if mapped_field == field_name and col in row:
                    value = row[col]
                    return str(value).strip() if value else None
            return None

        # Required fields
        references = get_field('reference')
        if not references:
            return None  # Skip rows without references

        # Parse quantity
        qty_str = get_field('quantity')
        try:
            quantity = int(qty_str) if qty_str else 1
        except ValueError:
            # Try to extract number from string like "10 pcs"
            qty_match = re.search(r'(\d+)', qty_str or '')
            quantity = int(qty_match.group(1)) if qty_match else 1

        # Parse unit price
        price_str = get_field('unit_price')
        unit_price = None
        if price_str:
            try:
                # Remove currency symbols and parse
                price_clean = re.sub(r'[^\d.]', '', price_str)
                unit_price = float(price_clean) if price_clean else None
            except ValueError:
                pass

        # Infer component type from reference
        component_type = self._infer_component_type(references)

        # Build item
        item = ParsedBOMItem(
            line_number=line_number,
            references=references,
            quantity=quantity,
            value=get_field('value'),
            part_number=get_field('part_number'),
            manufacturer=get_field('manufacturer'),
            description=get_field('description'),
            footprint=get_field('footprint'),
            supplier=get_field('supplier'),
            supplier_part_number=get_field('supplier_part_number'),
            unit_price=unit_price,
            component_type=component_type,
        )

        # Store unmapped columns as properties
        for col, value in row.items():
            if col not in column_map and value:
                item.properties[col] = str(value)

        return item

    def _infer_component_type(self, references: str) -> Optional[str]:
        """Infer component type from reference designators.

        Args:
            references: Comma-separated references (e.g., "R1,R2,R3")

        Returns:
            Component type string
        """
        # Get first reference
        first_ref = references.split(',')[0].strip()

        # Extract prefix (letters before numbers)
        prefix_match = re.match(r'^([A-Z]+)', first_ref.upper())
        if not prefix_match:
            return None

        prefix = prefix_match.group(1)

        # Map common prefixes to component types
        type_map = {
            'R': 'resistor',
            'C': 'capacitor',
            'L': 'inductor',
            'U': 'ic',
            'IC': 'ic',
            'Q': 'transistor',
            'D': 'diode',
            'LED': 'led',
            'J': 'connector',
            'P': 'connector',
            'SW': 'switch',
            'K': 'relay',
            'F': 'fuse',
            'T': 'transformer',
            'Y': 'crystal',
            'X': 'crystal',
            'BT': 'battery',
            'TP': 'test_point',
            'FB': 'ferrite_bead',
            'RN': 'resistor_network',
        }

        return type_map.get(prefix)
